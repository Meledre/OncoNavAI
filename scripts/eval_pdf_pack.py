#!/usr/bin/env python3
from __future__ import annotations

import argparse
import base64
import concurrent.futures
import json
import os
import threading
import time
import urllib.error
import urllib.request
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone
from http.cookies import SimpleCookie
from pathlib import Path
from typing import Any
from zipfile import ZipFile

PILOT_CASE_IDS: list[str] = [
    "PDF-0049",
    "PDF-0086",
    "PDF-0130",
    "PDF-0012",
    "PDF-0003",
    "PDF-0007",
    "PDF-0014",
    "PDF-0002",
    "PDF-0010",
    "PDF-0017",
    "PDF-0011",
    "PDF-0001",
    "PDF-0004",
    "PDF-0019",
    "PDF-0021",
    "PDF-0030",
    "PDF-0040",
    "PDF-0043",
    "PDF-0016",
    "PDF-0015",
]

REQUIRED_XLSX_SHEETS = {"Summary", "FlaggedCases", "CaseList"}


@dataclass(frozen=True)
class ManifestEntry:
    case_id: str
    pdf_file: str
    doctor_expected: str
    patient_expected: str
    request_id: str
    schema_version: str
    query_type: str
    nosology_key: str
    stage_group: str
    setting: str
    line: str
    source_set: str


class SmokeApiClient:
    def __init__(self, *, base_url: str, auth_mode: str, timeout_sec: int) -> None:
        self.base_url = base_url.rstrip("/")
        self.auth_mode = auth_mode
        self.timeout_sec = timeout_sec
        self._cookie_cache: dict[str, str] = {}

    def _request_json(
        self,
        *,
        method: str,
        path: str,
        role: str,
        payload: dict[str, Any] | None = None,
        extra_headers: dict[str, str] | None = None,
    ) -> tuple[int, dict[str, Any]]:
        body = json.dumps(payload).encode("utf-8") if payload is not None else None
        headers = {"content-type": "application/json"}
        headers.update(self._role_headers(role))
        if extra_headers:
            headers.update(extra_headers)
        url = f"{self.base_url}{path}"

        status, text = self._http_call(
            method=method,
            url=url,
            headers=headers,
            body=body,
            timeout=max(self.timeout_sec, 300) if path in {"/api/analyze", "/api/patient/analyze"} else self.timeout_sec,
        )
        if status in {401, 403}:
            self._cookie_cache.pop(role, None)
            headers = {"content-type": "application/json"}
            headers.update(self._role_headers(role))
            if extra_headers:
                headers.update(extra_headers)
            status, text = self._http_call(
                method=method,
                url=url,
                headers=headers,
                body=body,
                timeout=max(self.timeout_sec, 300)
                if path in {"/api/analyze", "/api/patient/analyze"}
                else self.timeout_sec,
            )

        try:
            data = json.loads(text) if text else {}
        except json.JSONDecodeError:
            data = {"raw": text}
        if not isinstance(data, dict):
            data = {"payload": data}
        return status, data

    def _session_login(self, payload: dict[str, Any]) -> tuple[int, str, list[str]]:
        body = json.dumps(payload).encode("utf-8")
        status, text, headers = self._http_call_with_headers(
            method="POST",
            url=f"{self.base_url}/api/session/login",
            headers={"content-type": "application/json"},
            body=body,
            timeout=self.timeout_sec,
        )
        set_cookies = [value for key, value in headers if key.lower() == "set-cookie"]
        return status, text, set_cookies

    def _credentials_from_env(self, role: str) -> tuple[str, str]:
        prefix = role.upper()
        username = str(
            os.environ.get(f"ONCOAI_{prefix}_USERNAME")
            or os.environ.get(f"ONCO_SMOKE_{prefix}_USERNAME")
            or ""
        ).strip()
        password = str(
            os.environ.get(f"ONCOAI_{prefix}_PASSWORD")
            or os.environ.get(f"ONCO_SMOKE_{prefix}_PASSWORD")
            or ""
        )
        return username, password

    def _cookie_for_role(self, role: str) -> str:
        cached = self._cookie_cache.get(role)
        if cached:
            return cached

        username, password = self._credentials_from_env(role)
        normalized_mode = self.auth_mode.strip().lower()
        if normalized_mode not in {"auto", "demo", "credentials"}:
            raise RuntimeError(f"unsupported auth mode: {self.auth_mode}")

        if normalized_mode == "credentials" or (normalized_mode == "auto" and username and password):
            if not username or not password:
                raise RuntimeError(
                    f"credentials auth selected but missing username/password for role={role}; "
                    f"expected ONCOAI_{role.upper()}_USERNAME/PASSWORD or ONCO_SMOKE_*"
                )
            login_payload: dict[str, Any] = {"username": username, "password": password}
        else:
            login_payload = {"role": role}

        status, response_text, set_cookie_headers = self._session_login(login_payload)
        cookie = SimpleCookie()
        for raw_cookie in set_cookie_headers:
            cookie.load(raw_cookie)

        required = ["session_access", "session_access_sig", "session_refresh", "session_refresh_sig"]
        missing = [name for name in required if cookie.get(name) is None]
        if missing:
            raise RuntimeError(
                "failed to bootstrap session cookies "
                f"for role={role}; status={status}; missing={missing}; response={response_text!r}"
            )

        cookie_header = "; ".join(f"{name}={cookie[name].value}" for name in required)
        self._cookie_cache[role] = cookie_header
        return cookie_header

    def _role_headers(self, role: str) -> dict[str, str]:
        return {"cookie": self._cookie_for_role(role)}

    @staticmethod
    def _http_call(
        *,
        method: str,
        url: str,
        headers: dict[str, str],
        body: bytes | None,
        timeout: int,
    ) -> tuple[int, str]:
        status, text, _ = SmokeApiClient._http_call_with_headers(
            method=method,
            url=url,
            headers=headers,
            body=body,
            timeout=timeout,
        )
        return status, text

    @staticmethod
    def _http_call_with_headers(
        *,
        method: str,
        url: str,
        headers: dict[str, str],
        body: bytes | None,
        timeout: int,
    ) -> tuple[int, str, list[tuple[str, str]]]:
        request = urllib.request.Request(url, method=method, data=body, headers=headers)
        try:
            with urllib.request.urlopen(request, timeout=timeout) as response:
                return (
                    response.status,
                    response.read().decode("utf-8"),
                    [(str(k), str(v)) for k, v in response.headers.items()],
                )
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8") if exc.fp else ""
            return (
                exc.code,
                detail,
                [(str(k), str(v)) for k, v in (exc.headers.items() if exc.headers else [])],
            )


def _require_ok(status: int, payload: dict[str, Any], action: str) -> None:
    if 200 <= status < 300:
        return
    raise RuntimeError(f"{action} failed: HTTP {status}, payload={payload}")


def _load_json(data: bytes, *, label: str) -> dict[str, Any]:
    try:
        obj = json.loads(data.decode("utf-8"))
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"invalid json for {label}") from exc
    if not isinstance(obj, dict):
        raise RuntimeError(f"json payload for {label} must be an object")
    return obj


def _resolve_pack_root(zf: ZipFile) -> str:
    names = zf.namelist()
    manifest_candidates = [name for name in names if name.endswith("/manifest.jsonl")]
    if len(manifest_candidates) != 1:
        raise RuntimeError(
            "zip must contain exactly one */manifest.jsonl; "
            f"found={manifest_candidates}"
        )
    return manifest_candidates[0].rsplit("/manifest.jsonl", 1)[0]


def _read_manifest(zf: ZipFile, *, root: str) -> list[ManifestEntry]:
    manifest_name = f"{root}/manifest.jsonl"
    try:
        lines = zf.read(manifest_name).decode("utf-8").splitlines()
    except KeyError as exc:
        raise RuntimeError(f"missing manifest: {manifest_name}") from exc

    entries: list[ManifestEntry] = []
    for idx, line in enumerate(lines, start=1):
        if not line.strip():
            continue
        try:
            row = json.loads(line)
        except json.JSONDecodeError as exc:
            raise RuntimeError(f"manifest line #{idx} is invalid json") from exc
        if not isinstance(row, dict):
            raise RuntimeError(f"manifest line #{idx} must be object")

        case_id = str(row.get("case_id") or "").strip()
        pdf_file = str(row.get("pdf_file") or "").strip()
        doctor_expected = str(row.get("doctor_expected") or "").strip()
        patient_expected = str(row.get("patient_expected") or "").strip()
        request_id = str(row.get("request_id") or "").strip()
        schema_version = str(row.get("schema_version") or "").strip()
        query_type = str(row.get("query_type") or "NEXT_STEPS").strip() or "NEXT_STEPS"

        if not case_id or not pdf_file or not doctor_expected or not patient_expected:
            raise RuntimeError(f"manifest line #{idx} missing required fields case_id/pdf_file/doctor_expected/patient_expected")

        entries.append(
            ManifestEntry(
                case_id=case_id,
                pdf_file=pdf_file,
                doctor_expected=doctor_expected,
                patient_expected=patient_expected,
                request_id=request_id,
                schema_version=schema_version,
                query_type=query_type,
                nosology_key=str(row.get("nosology_key") or "").strip(),
                stage_group=str(row.get("stage_group") or "").strip(),
                setting=str(row.get("setting") or "").strip(),
                line=str(row.get("line") or "").strip(),
                source_set=str(row.get("source_set") or "").strip(),
            )
        )

    if not entries:
        raise RuntimeError("manifest has no cases")
    return entries


def _validate_pack_structure(zf: ZipFile, *, root: str, entries: list[ManifestEntry]) -> dict[str, int]:
    names = set(zf.namelist())
    missing: list[str] = []
    pdf_count = 0
    doctor_expected_count = 0
    patient_expected_count = 0

    for entry in entries:
        pdf_member = f"{root}/{entry.pdf_file}"
        doctor_member = f"{root}/{entry.doctor_expected}"
        patient_member = f"{root}/{entry.patient_expected}"
        if pdf_member not in names:
            missing.append(pdf_member)
        else:
            pdf_count += 1
        if doctor_member not in names:
            missing.append(doctor_member)
        else:
            doctor_expected_count += 1
        if patient_member not in names:
            missing.append(patient_member)
        else:
            patient_expected_count += 1

    if missing:
        preview = ", ".join(missing[:10])
        raise RuntimeError(f"zip is missing required members ({len(missing)}): {preview}")

    return {
        "cases": len(entries),
        "pdf": pdf_count,
        "doctor_expected": doctor_expected_count,
        "patient_expected": patient_expected_count,
    }


def _validate_xlsx_report(path: Path) -> dict[str, Any]:
    if not path.exists():
        raise RuntimeError(f"xlsx report not found: {path}")
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError("openpyxl is required to validate xlsx report") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_names = set(wb.sheetnames)
    missing = sorted(REQUIRED_XLSX_SHEETS.difference(sheet_names))
    if missing:
        raise RuntimeError(f"xlsx report missing required sheets: {missing}")

    metadata: dict[str, Any] = {"path": str(path), "sheets": list(wb.sheetnames)}
    if "CaseList" in wb.sheetnames:
        metadata["case_list_rows"] = int(wb["CaseList"].max_row or 0)
    if "FlaggedCases" in wb.sheetnames:
        metadata["flagged_rows"] = int(wb["FlaggedCases"].max_row or 0)
    return metadata


def _parse_case_list(value: str) -> list[str]:
    token = value.strip()
    if not token:
        return []
    path = Path(token)
    if path.exists() and path.is_file():
        case_ids = [line.strip() for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]
        return case_ids
    return [item.strip() for item in token.split(",") if item.strip()]


def _select_cases(
    *,
    entries: list[ManifestEntry],
    sample_mode: str,
    case_list: list[str],
) -> list[ManifestEntry]:
    by_id = {entry.case_id: entry for entry in entries}

    if case_list:
        missing = [case_id for case_id in case_list if case_id not in by_id]
        if missing:
            raise RuntimeError(f"case-list includes unknown case_id values: {missing}")
        return [by_id[case_id] for case_id in case_list]

    if sample_mode == "full":
        return sorted(entries, key=lambda item: item.case_id)

    if sample_mode == "pilot":
        missing = [case_id for case_id in PILOT_CASE_IDS if case_id not in by_id]
        if missing:
            raise RuntimeError(f"pilot list has missing case_id values in manifest: {missing}")
        return [by_id[case_id] for case_id in PILOT_CASE_IDS]

    raise RuntimeError(f"unsupported sample-mode: {sample_mode}")


def _extract_patient_summary(payload: dict[str, Any]) -> str:
    patient_explain = payload.get("patient_explain") if isinstance(payload.get("patient_explain"), dict) else {}
    return str(patient_explain.get("summary") or patient_explain.get("summary_plain") or "").strip()


def _collect_pack_citation_metrics(analyze_payload: dict[str, Any]) -> tuple[int, list[str], int]:
    doctor_report = analyze_payload.get("doctor_report") if isinstance(analyze_payload.get("doctor_report"), dict) else {}
    issues = doctor_report.get("issues") if isinstance(doctor_report.get("issues"), list) else []
    citations = doctor_report.get("citations") if isinstance(doctor_report.get("citations"), list) else []

    issue_count = len(issues)
    citation_ids: set[str] = set()
    sources_used: set[str] = set()

    for issue in issues:
        if not isinstance(issue, dict):
            continue
        for citation_id in issue.get("citation_ids") if isinstance(issue.get("citation_ids"), list) else []:
            normalized = str(citation_id).strip()
            if normalized:
                citation_ids.add(normalized)

    citations_by_id: dict[str, dict[str, Any]] = {}
    for citation in citations:
        if not isinstance(citation, dict):
            continue
        citation_id = str(citation.get("citation_id") or "").strip()
        if citation_id:
            citations_by_id[citation_id] = citation
        source_id = str(citation.get("source_id") or "").strip()
        if source_id:
            sources_used.add(source_id)

    for citation_id in citation_ids:
        source_id = str((citations_by_id.get(citation_id) or {}).get("source_id") or "").strip()
        if source_id:
            sources_used.add(source_id)

    citation_count = len(citation_ids) if citation_ids else len(citations)

    patient_explain = analyze_payload.get("patient_explain") if isinstance(analyze_payload.get("patient_explain"), dict) else {}
    for source in patient_explain.get("sources_used") if isinstance(patient_explain.get("sources_used"), list) else []:
        normalized = str(source).strip()
        if normalized:
            sources_used.add(normalized)

    return citation_count, sorted(sources_used), issue_count


def _top_level_diff(expected: dict[str, Any], actual: dict[str, Any]) -> dict[str, Any]:
    expected_keys = set(expected.keys())
    actual_keys = set(actual.keys())
    return {
        "missing_keys": sorted(expected_keys.difference(actual_keys)),
        "extra_keys": sorted(actual_keys.difference(expected_keys)),
        "schema_version_expected": expected.get("schema_version"),
        "schema_version_actual": actual.get("schema_version"),
    }


def _build_expected_diff(
    *,
    case_id: str,
    expected_doctor: dict[str, Any],
    expected_patient: dict[str, Any],
    actual_doctor: dict[str, Any],
    actual_patient: dict[str, Any],
    gate_failures: list[str],
) -> dict[str, Any]:
    return {
        "case_id": case_id,
        "doctor": _top_level_diff(expected_doctor, actual_doctor),
        "patient": _top_level_diff(expected_patient, actual_patient),
        "doctor_citations_expected": len(expected_doctor.get("citations") or []),
        "doctor_citations_actual": len(actual_doctor.get("citations") or []),
        "doctor_issues_expected": len(expected_doctor.get("issues") or []),
        "doctor_issues_actual": len(actual_doctor.get("issues") or []),
        "patient_key_points_expected": len(expected_patient.get("key_points") or []),
        "patient_key_points_actual": len(actual_patient.get("key_points") or []),
        "gate_failures": gate_failures,
    }


def _read_case_bundle(
    *,
    entry: ManifestEntry,
    root: str,
    zf: ZipFile | None,
    zip_path: Path | None,
) -> tuple[bytes, dict[str, Any], dict[str, Any]]:
    pdf_member = f"{root}/{entry.pdf_file}"
    doctor_expected_member = f"{root}/{entry.doctor_expected}"
    patient_expected_member = f"{root}/{entry.patient_expected}"

    if zf is not None:
        pdf_bytes = zf.read(pdf_member)
        expected_doctor = _load_json(zf.read(doctor_expected_member), label=f"{entry.case_id}.doctor")
        expected_patient = _load_json(zf.read(patient_expected_member), label=f"{entry.case_id}.patient")
        return pdf_bytes, expected_doctor, expected_patient

    if zip_path is None:
        raise RuntimeError("zip_path is required when zf is not provided")

    with ZipFile(zip_path) as local_zf:
        pdf_bytes = local_zf.read(pdf_member)
        expected_doctor = _load_json(local_zf.read(doctor_expected_member), label=f"{entry.case_id}.doctor")
        expected_patient = _load_json(local_zf.read(patient_expected_member), label=f"{entry.case_id}.patient")
    return pdf_bytes, expected_doctor, expected_patient


def _process_case(
    *,
    entry: ManifestEntry,
    root: str,
    client: SmokeApiClient,
    schema_version: str,
    zf: ZipFile | None = None,
    zip_path: Path | None = None,
) -> tuple[dict[str, Any], dict[str, Any]]:
    started_at = time.monotonic()
    gate_failures: list[str] = []

    pdf_bytes, expected_doctor, expected_patient = _read_case_bundle(
        entry=entry,
        root=root,
        zf=zf,
        zip_path=zip_path,
    )
    content_base64 = base64.b64encode(pdf_bytes).decode("utf-8")

    import_payload = {
        "filename": Path(entry.pdf_file).name,
        "content_base64": content_base64,
        "mime_type": "application/pdf",
        "data_mode": "DEID",
    }
    import_status, import_body = client._request_json(
        method="POST",
        path="/api/case/import-file",
        role="clinician",
        payload=import_payload,
    )
    if not (200 <= import_status < 300):
        gate_failures.append(f"import_file_failed status={import_status}")
    case_id = str(import_body.get("case_id") or "").strip()
    import_run_id = str(import_body.get("import_run_id") or "").strip()
    if not case_id:
        gate_failures.append("missing_case_id")
    if not import_run_id:
        gate_failures.append("missing_import_run_id")

    analyze_body: dict[str, Any] = {}
    patient_body: dict[str, Any] = {}
    analyze_status = 0
    patient_status = 0

    if case_id:
        analyze_request = {
            "schema_version": schema_version,
            "request_id": str(uuid.uuid4()),
            "query_type": entry.query_type if entry.query_type in {"NEXT_STEPS", "CHECK_LAST_TREATMENT"} else "NEXT_STEPS",
            "sources": {"mode": "AUTO", "source_ids": ["minzdrav", "russco"]},
            "language": "ru",
            "case": {"case_id": case_id},
            "options": {"strict_evidence": True, "max_chunks": 40, "max_citations": 40, "timeout_ms": 120000},
        }
        analyze_status, analyze_body = client._request_json(
            method="POST",
            path="/api/analyze",
            role="clinician",
            payload=analyze_request,
            extra_headers={"x-client-id": f"pdf-pack-{entry.case_id.lower()}"},
        )
        if not (200 <= analyze_status < 300):
            gate_failures.append(f"analyze_failed status={analyze_status}")

        patient_request = {
            "filename": Path(entry.pdf_file).name,
            "content_base64": content_base64,
            "mime_type": "application/pdf",
            "request_id": str(uuid.uuid4()),
            "query_type": entry.query_type if entry.query_type in {"NEXT_STEPS", "CHECK_LAST_TREATMENT"} else "NEXT_STEPS",
            "sources": {"mode": "AUTO", "source_ids": ["minzdrav", "russco"]},
            "language": "ru",
        }
        patient_status, patient_body = client._request_json(
            method="POST",
            path="/api/patient/analyze",
            role="patient",
            payload=patient_request,
        )
        if not (200 <= patient_status < 300):
            gate_failures.append(f"patient_analyze_failed status={patient_status}")

    doctor_report = analyze_body.get("doctor_report") if isinstance(analyze_body.get("doctor_report"), dict) else {}
    patient_explain = analyze_body.get("patient_explain") if isinstance(analyze_body.get("patient_explain"), dict) else {}
    patient_payload_explain = patient_body.get("patient_explain") if isinstance(patient_body.get("patient_explain"), dict) else {}

    if not doctor_report:
        gate_failures.append("missing_doctor_report")
    if not patient_explain:
        gate_failures.append("missing_doctor_patient_explain")
    if not patient_payload_explain:
        gate_failures.append("missing_patient_patient_explain")
    if "doctor_report" in patient_body:
        gate_failures.append("patient_endpoint_leaked_doctor_report")

    citations_count, sources_used, issues_count = _collect_pack_citation_metrics(analyze_body)
    if issues_count < 1:
        gate_failures.append(f"issues_count_lt_1 got={issues_count}")
    if citations_count < 1:
        gate_failures.append(f"citations_count_lt_1 got={citations_count}")
    if not _extract_patient_summary(analyze_body):
        gate_failures.append("doctor_patient_summary_empty")
    patient_summary = str(
        patient_payload_explain.get("summary") or patient_payload_explain.get("summary_plain") or ""
    ).strip()
    if not patient_summary:
        gate_failures.append("patient_summary_empty")

    report_id = str(doctor_report.get("report_id") or "").strip()
    if not report_id:
        gate_failures.append("missing_doctor_report_id")

    diff_item = _build_expected_diff(
        case_id=entry.case_id,
        expected_doctor=expected_doctor,
        expected_patient=expected_patient,
        actual_doctor=doctor_report,
        actual_patient=patient_payload_explain,
        gate_failures=gate_failures,
    )

    duration_sec = round(time.monotonic() - started_at, 3)
    case_result = {
        "case_id": entry.case_id,
        "manifest": {
            "nosology_key": entry.nosology_key,
            "stage_group": entry.stage_group,
            "setting": entry.setting,
            "line": entry.line,
            "query_type": entry.query_type,
            "source_set": entry.source_set,
        },
        "runtime": {
            "import_status": import_status,
            "analyze_status": analyze_status,
            "patient_status": patient_status,
            "case_id": case_id or None,
            "import_run_id": import_run_id or None,
            "report_id": report_id or None,
            "issues_count": issues_count,
            "citations_count": citations_count,
            "sources_used": sources_used,
            "duration_sec": duration_sec,
        },
        "gate_failures": gate_failures,
        "passed": not gate_failures,
    }
    return case_result, diff_item


def _exception_result(entry: ManifestEntry, exc: Exception) -> tuple[dict[str, Any], dict[str, Any]]:
    reason = str(exc).strip() or repr(exc)
    case_row = {
        "case_id": entry.case_id,
        "manifest": {
            "nosology_key": entry.nosology_key,
            "stage_group": entry.stage_group,
            "setting": entry.setting,
            "line": entry.line,
            "query_type": entry.query_type,
            "source_set": entry.source_set,
        },
        "runtime": {
            "import_status": 0,
            "analyze_status": 0,
            "patient_status": 0,
            "case_id": None,
            "import_run_id": None,
            "report_id": None,
            "issues_count": 0,
            "citations_count": 0,
            "sources_used": [],
            "duration_sec": 0.0,
        },
        "gate_failures": [f"runtime_exception:{reason}"],
        "passed": False,
    }
    diff_item = {
        "case_id": entry.case_id,
        "error": reason,
        "gate_failures": case_row["gate_failures"],
    }
    return case_row, diff_item


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8") as fh:
        for row in rows:
            fh.write(json.dumps(row, ensure_ascii=False) + "\n")


def _render_markdown_report(*, summary: dict[str, Any], results: list[dict[str, Any]]) -> str:
    fail_rows = [item for item in results if not bool(item.get("passed"))]
    lines = [
        "# PDF Pack Gate Report",
        "",
        f"- Run ID: `{summary['run_id']}`",
        f"- Mode: `{summary['sample_mode']}`",
        f"- Dataset status: `shadow_candidate`",
        f"- Cases total: `{summary['cases_total']}`",
        f"- Passed: `{summary['cases_passed']}`",
        f"- Failed: `{summary['cases_failed']}`",
        "",
    ]
    if fail_rows:
        lines.append("## Failed Cases")
        lines.append("")
        for row in fail_rows:
            case_id = str(row.get("case_id") or "")
            failures = row.get("gate_failures") if isinstance(row.get("gate_failures"), list) else []
            lines.append(f"- `{case_id}`: {', '.join(str(item) for item in failures)}")
    else:
        lines.append("## Result")
        lines.append("")
        lines.append("All selected cases passed hard quality gates.")
    lines.append("")
    return "\n".join(lines)


def _default_output_dir() -> Path:
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    root = Path(__file__).resolve().parents[1]
    return root / "reports" / "metrics" / f"pdf_pack_v8_smooth_{ts}"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Evaluate synthetic PDF pack against release-like quality gates")
    parser.add_argument("--zip", required=True, help="Path to synthetic_pdf_cases_*.zip")
    parser.add_argument("--xlsx", default="", help="Optional xlsx summary report path for intake validation")
    parser.add_argument("--base-url", default="http://localhost:3000")
    parser.add_argument("--auth-mode", choices=["auto", "demo", "credentials"], default="demo")
    parser.add_argument("--schema-version", choices=["0.2"], default="0.2")
    parser.add_argument("--sample-mode", choices=["pilot", "full"], default="pilot")
    parser.add_argument("--case-list", default="", help="Comma-separated case ids or path to file with one case_id per line")
    parser.add_argument("--out-dir", default="", help="Output directory path (default reports/metrics/pdf_pack_v8_smooth_<ts>)")
    parser.add_argument("--http-timeout", type=int, default=180)
    parser.add_argument("--workers", type=int, default=1, help="Parallel case workers (1 = sequential)")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.workers < 1:
        raise RuntimeError("--workers must be >= 1")
    zip_path = Path(args.zip).expanduser().resolve()
    if not zip_path.exists():
        raise RuntimeError(f"zip not found: {zip_path}")

    out_dir = Path(args.out_dir).expanduser().resolve() if args.out_dir else _default_output_dir()
    out_dir.mkdir(parents=True, exist_ok=True)

    xlsx_meta: dict[str, Any] = {}
    if str(args.xlsx).strip():
        xlsx_meta = _validate_xlsx_report(Path(args.xlsx).expanduser().resolve())

    with ZipFile(zip_path) as zf:
        root = _resolve_pack_root(zf)
        entries = _read_manifest(zf, root=root)
        pack_counts = _validate_pack_structure(zf, root=root, entries=entries)
        selected_case_ids = _parse_case_list(args.case_list)
        selected_entries = _select_cases(entries=entries, sample_mode=args.sample_mode, case_list=selected_case_ids)

        per_case_rows: list[dict[str, Any]] = []
        diff_rows: list[dict[str, Any]] = []
        failures: list[dict[str, Any]] = []
        started_wall = time.monotonic()

        if args.workers == 1:
            client = SmokeApiClient(base_url=args.base_url, auth_mode=args.auth_mode, timeout_sec=max(10, args.http_timeout))
            for idx, entry in enumerate(selected_entries, start=1):
                try:
                    case_row, diff_row = _process_case(
                        entry=entry,
                        zf=zf,
                        root=root,
                        client=client,
                        schema_version=args.schema_version,
                    )
                except Exception as exc:  # noqa: BLE001
                    case_row, diff_row = _exception_result(entry, exc)
                per_case_rows.append(case_row)
                diff_rows.append(diff_row)
                if not bool(case_row.get("passed")):
                    failures.append(case_row)
                duration = case_row.get("runtime", {}).get("duration_sec") if isinstance(case_row.get("runtime"), dict) else None
                print(
                    f"[{idx}/{len(selected_entries)}] {entry.case_id} "
                    f"{'PASS' if case_row.get('passed') else 'FAIL'} "
                    f"duration={duration}s",
                    flush=True,
                )
        else:
            thread_state = threading.local()

            def _worker(entry: ManifestEntry) -> tuple[str, dict[str, Any], dict[str, Any]]:
                client = getattr(thread_state, "client", None)
                if client is None:
                    client = SmokeApiClient(
                        base_url=args.base_url,
                        auth_mode=args.auth_mode,
                        timeout_sec=max(10, args.http_timeout),
                    )
                    thread_state.client = client
                try:
                    case_row, diff_row = _process_case(
                        entry=entry,
                        zip_path=zip_path,
                        root=root,
                        client=client,
                        schema_version=args.schema_version,
                    )
                except Exception as exc:  # noqa: BLE001
                    case_row, diff_row = _exception_result(entry, exc)
                return entry.case_id, case_row, diff_row

            ordered_case_ids = [entry.case_id for entry in selected_entries]
            rows_by_case_id: dict[str, dict[str, Any]] = {}
            diffs_by_case_id: dict[str, dict[str, Any]] = {}
            completed = 0

            with concurrent.futures.ThreadPoolExecutor(max_workers=args.workers) as executor:
                future_map = {executor.submit(_worker, entry): entry for entry in selected_entries}
                for future in concurrent.futures.as_completed(future_map):
                    entry = future_map[future]
                    case_id, case_row, diff_row = future.result()
                    rows_by_case_id[case_id] = case_row
                    diffs_by_case_id[case_id] = diff_row
                    completed += 1
                    duration = case_row.get("runtime", {}).get("duration_sec") if isinstance(case_row.get("runtime"), dict) else None
                    print(
                        f"[{completed}/{len(selected_entries)}] {entry.case_id} "
                        f"{'PASS' if case_row.get('passed') else 'FAIL'} "
                        f"duration={duration}s",
                        flush=True,
                    )

            for case_id in ordered_case_ids:
                case_row = rows_by_case_id[case_id]
                diff_row = diffs_by_case_id[case_id]
                per_case_rows.append(case_row)
                diff_rows.append(diff_row)
                if not bool(case_row.get("passed")):
                    failures.append(case_row)

        wall_duration_sec = round(time.monotonic() - started_wall, 3)

    run_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    case_runtime_values = [
        float(item.get("runtime", {}).get("duration_sec") or 0.0)
        for item in per_case_rows
        if isinstance(item.get("runtime"), dict)
    ]
    avg_case_runtime_sec = round(sum(case_runtime_values) / len(case_runtime_values), 3) if case_runtime_values else 0.0
    effective_case_throughput_per_min = round((len(per_case_rows) / wall_duration_sec) * 60, 3) if wall_duration_sec > 0 else 0.0
    summary = {
        "run_id": run_id,
        "dataset_status": "shadow_candidate",
        "zip_path": str(zip_path),
        "sample_mode": args.sample_mode,
        "schema_version": args.schema_version,
        "auth_mode": args.auth_mode,
        "base_url": args.base_url,
        "counts": pack_counts,
        "cases_total": len(per_case_rows),
        "cases_passed": len(per_case_rows) - len(failures),
        "cases_failed": len(failures),
        "workers": args.workers,
        "wall_duration_sec": wall_duration_sec,
        "avg_case_runtime_sec": avg_case_runtime_sec,
        "effective_case_throughput_per_min": effective_case_throughput_per_min,
        "pilot_case_ids": PILOT_CASE_IDS,
        "selected_case_ids": [entry.case_id for entry in selected_entries],
        "xlsx": xlsx_meta,
        "artifacts_dir": str(out_dir),
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }

    _write_json(out_dir / "summary.json", summary)
    _write_jsonl(out_dir / "per_case.jsonl", per_case_rows)
    _write_jsonl(out_dir / "failures.jsonl", failures)
    _write_jsonl(out_dir / "expected_diff.jsonl", diff_rows)
    (out_dir / "pilot_case_ids.txt").write_text("\n".join(PILOT_CASE_IDS) + "\n", encoding="utf-8")
    (out_dir / "gates_report.md").write_text(
        _render_markdown_report(summary=summary, results=per_case_rows),
        encoding="utf-8",
    )

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    if failures:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
